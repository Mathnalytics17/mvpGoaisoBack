from itertools import permutations
import random

from django.shortcuts import get_object_or_404
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from toon_format import decode
from collections import Counter
from apps.results.api.models.index import (
    Evaluation,
    EvaluationCriterion,
    PromptRun,
    RankingItem,
    RankingSummary,
)
from django.db import transaction
from apps.results.api.serializers.index import (
    EvaluationSerializer,
    EvaluationCreateSerializer,
)

from apps.results.services.prompts import (
    prompt_toon_phase1,
    prompt_toon_phase2,
)

from apps.results.services.scoring import compute_brand_summary
from apps.results.services.parse_ranking import parse_ranking

from apps.results.utils.open_ai_client import completion_with_web_search



from django.http import HttpResponse
from django.utils.timezone import now

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny
from rest_framework.filters import SearchFilter, OrderingFilter

from django_filters.rest_framework import DjangoFilterBackend

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from apps.results.api.models.index import InformeDataUsers
from apps.results.api.serializers.index import InformeDataUsersCreateSerializer, InformeDataUsersListSerializer


from apps.results.api.models.index import Evaluation
from apps.results.services.report import build_report



import os
import tempfile
import subprocess

from django.conf import settings
from django.http import FileResponse, HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny

# ✅ helper para seleccionar 5 permutaciones sin repetir el mismo inicio
def select_permutations_unique_start(permutations_list, count=5):
    selected = []
    used_first = set()

    random.shuffle(permutations_list)

    for perm in permutations_list:
        if perm[0] not in used_first:
            selected.append(perm)
            used_first.add(perm[0])
        if len(selected) == count:
            break

    for perm in permutations_list:
        if perm not in selected:
            selected.append(perm)
        if len(selected) == count:
            break

    return selected

class EvaluationCreateView(APIView):
    def post(self, request):
        serializer = EvaluationCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        country = serializer.validated_data.get("country", "").strip() or None
        location = serializer.validated_data.get("location", "").strip() or None

        evaluation = Evaluation.objects.create(
            product_type=serializer.validated_data["product_type"],
            status="PENDING",
            country=country,
            location=location,
        )

        for idx, c in enumerate(serializer.validated_data["criteria"], start=1):
            EvaluationCriterion.objects.create(
                evaluation=evaluation,
                name=c,
                order=idx,
            )

        return Response(
            {"uuid": str(evaluation.uuid), "status": evaluation.status},
            status=status.HTTP_201_CREATED,
        )


class EvaluationDetailView(APIView):
    """
    GET /api/results/<uuid>/
    """

    def get(self, request, uuid):
        evaluation = get_object_or_404(Evaluation, uuid=uuid)
        serializer = EvaluationSerializer(evaluation)
        return Response(serializer.data)


class EvaluationListView(APIView):
    """
    GET /api/results/
    Lista todas las evaluaciones
    """

    def get(self, request):
        evaluations = Evaluation.objects.all().order_by("-created_at")
        serializer = EvaluationSerializer(evaluations, many=True)
        return Response(serializer.data)


class JsonToToonView(APIView):
    """
    POST /api/results/json-to-toon/
    Body: JSON
    Response: TOON (string)
    """

    def post(self, request):
        try:
            toon_text = encode(request.data)
            return Response({"toon": toon_text}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                {"error": "No se pudo convertir a TOON", "details": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class RunEvaluationView(APIView):

    def post(self, request, uuid):

        # ==========================
        # ✅ LOCK POR UUID (evita dobles ejecuciones)
        # ==========================
        with transaction.atomic():
            evaluation = Evaluation.objects.select_for_update().get(uuid=uuid)

            # ✅ Si ya está corriendo -> no permitir doble ejecución
            if evaluation.status == "PROCESSING":
                return Response(
                    {"error": "Esta evaluación ya se está ejecutando"},
                    status=status.HTTP_409_CONFLICT,
                )

            # ✅ Reset seguro antes de correr
            evaluation.status = "PROCESSING"
            evaluation.completed_at = None
            evaluation.save()

            # ✅ Limpiar runs anteriores dentro del lock
            PromptRun.objects.filter(evaluation=evaluation).delete()
            RankingItem.objects.filter(prompt_run__evaluation=evaluation).delete()
            RankingSummary.objects.filter(evaluation=evaluation).delete()

        # ==========================
        # ✅ Ya salimos del lock, empieza el proceso real
        # ==========================

        try:
            criteria_qs = evaluation.criteria.all().order_by("order")
            criteria = [c.name for c in criteria_qs]

            if len(criteria) < 2:
                evaluation.status = "ERROR"
                evaluation.save()
                return Response(
                    {"error": "Se requieren mínimo 2 criterios"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # =========================
            # ✅ PHASE 1 (5 permutaciones)
            # =========================
            all_perms = list(permutations(criteria, len(criteria)))
            selected_perms = select_permutations_unique_start(all_perms, 5)

            for perm in selected_perms:
                criteria_str = ", ".join(perm)

                prompt = prompt_toon_phase1(
                    evaluation.product_type,
                    criteria_str,
                    country=evaluation.country,
                    location=evaluation.location
                )

                toon_text, sources = completion_with_web_search(prompt)

                try:
                    decoded = decode(toon_text)
                except Exception as e:
                    evaluation.status = "ERROR"
                    evaluation.save()
                    return Response(
                        {"error": "TOON inválido PHASE1", "toon": toon_text, "details": str(e)},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                parsed = parse_ranking(decoded)
                if not parsed or len(parsed) != 5:
                    evaluation.status = "ERROR"
                    evaluation.save()
                    return Response(
                        {"error": "ranking inválido PHASE1", "decoded": decoded, "toon": toon_text},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                run = PromptRun.objects.create(
                    evaluation=evaluation,
                    phase="PHASE1",
                    prompt_text=prompt,
                    response_raw=toon_text,
                    sources=sources,
                )

                for item in parsed:
                    RankingItem.objects.create(
                        prompt_run=run,
                        position=item["position"],
                        brand=item["brand"],
                        model=item["model"],
                        raw_text=item["raw_text"],
                    )

            compute_brand_summary(evaluation, phase="PHASE1")

            # =========================
            # ✅ PHASE 2 (5 prompts por criterio SIEMPRE)
            # =========================
            for criterion_obj in criteria_qs:
                for _ in range(5):

                    prompt = prompt_toon_phase2(
                        evaluation.product_type,
                        criterion_obj.name,
                        country=evaluation.country,
                        location=evaluation.location
                    )

                    toon_text, sources = completion_with_web_search(prompt)

                    try:
                        decoded = decode(toon_text)
                    except Exception as e:
                        evaluation.status = "ERROR"
                        evaluation.save()
                        return Response(
                            {"error": "TOON inválido PHASE2", "toon": toon_text, "details": str(e)},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    parsed = parse_ranking(decoded)
                    if not parsed or len(parsed) != 5:
                        evaluation.status = "ERROR"
                        evaluation.save()
                        return Response(
                            {"error": "ranking inválido PHASE2", "decoded": decoded, "toon": toon_text},
                            status=status.HTTP_400_BAD_REQUEST,
                        )

                    run = PromptRun.objects.create(
                        evaluation=evaluation,
                        phase="PHASE2",
                        criterion=criterion_obj,
                        prompt_text=prompt,
                        response_raw=toon_text,
                        sources=sources,
                    )

                    for item in parsed:
                        RankingItem.objects.create(
                            prompt_run=run,
                            position=item["position"],
                            brand=item["brand"],
                            model=item["model"],
                            raw_text=item["raw_text"],
                        )

                compute_brand_summary(evaluation, phase="PHASE2", criterion=criterion_obj)

            # ✅ SUCCESS
            evaluation.status = "SUCCESS"
            evaluation.completed_at = timezone.now()
            evaluation.save()

            return Response(
                {"status": evaluation.status, "uuid": str(evaluation.uuid)},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            # ✅ Cualquier fallo inesperado → marca ERROR
            evaluation.status = "ERROR"
            evaluation.save()
            return Response(
                {"error": "Error inesperado ejecutando evaluación", "details": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )



class EvaluationReportView(APIView):
    def get(self, request, uuid):
        evaluation = get_object_or_404(Evaluation, uuid=uuid)
        return Response(build_report(evaluation), status=status.HTTP_200_OK)


def apply_filters(request, qs):
    """
    Filtros soportados (query params):
      - search: busca en nombre/email/movil/evaluation_uuid
      - nombre, email, movil: icontains
      - uuid: filtro exacto por evaluation.uuid
      - ordering: id | -id | nombre | -nombre | email | -email
    """
    search = (request.GET.get("search") or "").strip()
    nombre = (request.GET.get("nombre") or "").strip()
    email = (request.GET.get("email") or "").strip()
    movil = (request.GET.get("movil") or "").strip()
    uuid = (request.GET.get("uuid") or "").strip()

    if uuid:
        qs = qs.filter(evaluation__uuid=uuid)

    if nombre:
        qs = qs.filter(nombre__icontains=nombre)
    if email:
        qs = qs.filter(email__icontains=email)
    if movil:
        qs = qs.filter(movil__icontains=movil)

    if search:
        # search global
        qs = qs.filter(
            nombre__icontains=search
        ) | qs.filter(
            email__icontains=search
        ) | qs.filter(
            movil__icontains=search
        ) | qs.filter(
            evaluation__uuid__icontains=search
        )

    ordering = (request.GET.get("ordering") or "-id").strip()
    allowed = {"id", "-id", "nombre", "-nombre", "email", "-email"}
    if ordering not in allowed:
        ordering = "-id"
    return qs.order_by(ordering)


class InformeDataUsersAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        LIST con paginación simple:
          GET /api/results/report/users/?page=1&page_size=20&search=...
        """
        qs = InformeDataUsers.objects.select_related("evaluation").all()

        qs = apply_filters(request, qs)

        # paginación
        try:
            page = int(request.GET.get("page", "1"))
        except ValueError:
            page = 1
        try:
            page_size = int(request.GET.get("page_size", "20"))
        except ValueError:
            page_size = 20

        page = max(1, page)
        page_size = min(max(1, page_size), 200)

        total = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        rows = qs[start:end]

        serializer = InformeDataUsersListSerializer(rows, many=True)
        return Response(
            {
                "count": total,
                "page": page,
                "page_size": page_size,
                "results": serializer.data,
            },
            status=status.HTTP_200_OK,
        )

    def post(self, request):
        """
        CREATE:
          POST /api/results/report/users/
          body: { uuid, nombre, email, movil? }
        """
        serializer = InformeDataUsersCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        lead = serializer.save()

        return Response(
            {
                "ok": True,
                "id": lead.id,
                "evaluation_uuid": str(lead.evaluation.uuid),
            },
            status=status.HTTP_201_CREATED,
        )


class InformeDataUsersExportAPIView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Excel:
          GET /api/results/report/users/export/?search=...&uuid=...
        """
        qs = InformeDataUsers.objects.select_related("evaluation").all()
        qs = apply_filters(request, qs)

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads Informe"

        headers = ["ID", "Evaluation UUID", "Nombre", "Email", "Móvil"]
        ws.append(headers)

        header_font = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = header_font
            c.alignment = Alignment(vertical="center")

        for row in qs.iterator():
            ws.append(
                [
                    row.id,
                    str(row.evaluation.uuid) if row.evaluation_id else "",
                    row.nombre,
                    row.email,
                    row.movil or "",
                ]
            )

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        widths = [8, 40, 28, 34, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        filename = f"leads_informe_{now().strftime('%Y%m%d_%H%M')}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp
class EvaluationReportPDFView(APIView):
    """
    GET /api/results/<uuid>/report/pdf/
    Genera PDF renderizando el frontend con Puppeteer.
    """
    permission_classes = [AllowAny]

    def get(self, request, uuid):
        # URL del frontend accesible desde backend
        front_base = getattr(settings, "FRONTEND_BASE_URL", "http://localhost:3000")
        report_url = f"{front_base}/results/{uuid}?pdf=1"

        script_path = os.path.join(settings.BASE_DIR, "apps","base","scripts", "render_report_pdf.js")

        # Archivo temporal
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        out_path = tmp.name
        tmp.close()

        try:
            # Ejecuta Node + Puppeteer
            completed = subprocess.run(
                ["node", script_path, report_url, out_path],
                capture_output=True,
                text=True,
                check=True,
            )

            filename = f"informe_goaiso_{uuid}.pdf"
            resp = FileResponse(open(out_path, "rb"), content_type="application/pdf")
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

        except subprocess.CalledProcessError as e:
            return HttpResponse(
                "Error generando PDF.\n\nSTDOUT:\n"
                + (e.stdout or "")
                + "\n\nSTDERR:\n"
                + (e.stderr or ""),
                status=500,
                content_type="text/plain",
            )
        finally:
            # (Opcional) borrar el archivo luego con un cleanup job
            # No lo borro aquí por seguridad con FileResponse en algunos entornos.
            pass