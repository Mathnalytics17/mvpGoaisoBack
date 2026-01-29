import os
import time
import logging
from datetime import datetime
from dotenv import load_dotenv
from openai import OpenAI
from toon_format import decode

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

load_dotenv()

# -------------------------
# Logger
# -------------------------
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler()
console_handler.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)

if not logger.handlers:
    logger.addHandler(console_handler)

# -------------------------
# OpenAI client
# -------------------------
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
DEFAULT_MODEL = "gpt-4o-mini"


# -------------------------
# Excel helpers
# -------------------------
DEFAULT_XLSX_PATH = "websearch_logs.xlsx"

HEADERS = [
    "timestamp",
    "model",
    "phase",
    "criterion",
    "evaluation_uuid",
    "attempt",
    "elapsed_seconds",
    "toon_valid",
    "prompt",
    "output_text",
    "sources",
]

def ensure_workbook(path: str) -> tuple[Workbook, any]:
    """Crea o abre el Excel y asegura header + estilos."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "logs"
        ws.append(HEADERS)

        # estilo header
        header_font = Font(bold=True)
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        # anchos razonables
        widths = {
            "A": 20,  # timestamp
            "B": 14,  # model
            "C": 10,  # phase
            "D": 18,  # criterion
            "E": 38,  # evaluation_uuid
            "F": 8,   # attempt
            "G": 14,  # elapsed
            "H": 10,  # valid
            "I": 70,  # prompt
            "J": 70,  # output
            "K": 70,  # sources
        }
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    return wb, ws

def append_log_row(
    xlsx_path: str,
    *,
    model: str,
    attempt: int,
    elapsed: float,
    toon_valid: bool,
    prompt: str,
    output_text: str,
    sources: list[str],
    phase: str = "",
    criterion: str = "",
    evaluation_uuid: str = "",
):
    wb, ws = ensure_workbook(xlsx_path)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sources_str = " | ".join(sources)  # una celda

    ws.append([
        ts,
        model,
        phase,
        criterion,
        evaluation_uuid,
        attempt,
        float(elapsed),
        "YES" if toon_valid else "NO",
        prompt,
        output_text,
        sources_str,
    ])

    # refresca autofiltro (por si crece)
    ws.auto_filter.ref = ws.dimensions

    wb.save(xlsx_path)


# -------------------------
# Core logic
# -------------------------
def _clean_output_text(text: str) -> str:
    text = (text or "").strip()
    text = text.replace("\n", " ").strip()
    text = text.replace("```", "").strip()
    return text

def _extract_sources(res) -> list[str]:
    sources = []
    try:
        for item in res.output:
            if getattr(item, "type", None) == "web_search_call":
                results = getattr(item, "results", None)
                if results:
                    for r in results:
                        url = r.get("url")
                        if url:
                            sources.append(url)
    except Exception as e:
        logger.warning(f"[SOURCES ERROR] {str(e)}")

    return list(dict.fromkeys(sources))[:10]

def _is_valid_toon_ranking5(output_text: str) -> bool:
    try:
        decoded = decode(output_text)
        return (
            "ranking" in decoded
            and isinstance(decoded["ranking"], list)
            and len(decoded["ranking"]) == 5
        )
    except Exception:
        return False


def completion_with_web_search(
    prompt: str,
    model: str = DEFAULT_MODEL,
    max_retries: int = 2,
    *,
    xlsx_path: str = DEFAULT_XLSX_PATH,
    phase: str = "",
    criterion: str = "",
    evaluation_uuid: str = "",
):
    """
    ✅ web_search + logs + retry
    ✅ guarda CADA intento en Excel (para pruebas)
    """

    last_output_text = ""
    last_sources: list[str] = []

    for attempt in range(1, max_retries + 2):
        start = time.time()

        logger.debug("=" * 60)
        logger.debug(f"[WEBSEARCH] Attempt {attempt}")
        logger.debug(f"[WEBSEARCH] Model: {model}")
        logger.debug(f"[PROMPT PREVIEW] {prompt[:250]}...")

        res = client.responses.create(
            model=model,
            input=prompt,
            tools=[{"type": "web_search", "search_context_size": "low"}],
        )

        elapsed = round(time.time() - start, 2)

        output_text = _clean_output_text(res.output_text)
        sources = _extract_sources(res)

        valid = _is_valid_toon_ranking5(output_text)

        logger.debug(f"[OUTPUT RAW] {output_text[:400]}...")
        logger.debug(f"[TIME] {elapsed}s")
        logger.debug(f"[SOURCES] {len(sources)} found")
        logger.debug(f"[TOON VALID] {valid}")

        # ✅ guarda SIEMPRE el intento en Excel
        append_log_row(
            xlsx_path,
            model=model,
            attempt=attempt,
            elapsed=elapsed,
            toon_valid=valid,
            prompt=prompt,
            output_text=output_text,
            sources=sources,
            phase=phase,
            criterion=criterion,
            evaluation_uuid=evaluation_uuid,
        )

        last_output_text = output_text
        last_sources = sources

        if valid:
            return output_text, sources

        logger.warning("[RETRYING] TOON inválido, intentando de nuevo...")
        time.sleep(1)

    logger.error("[FAILED] No se obtuvo TOON válido tras varios intentos")
    return last_output_text, last_sources
